# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./test.xlsx')

# Get sheet names
#sheetnames = wb.get_sheet_names()
#for i in sheetnames:
#print (i)

#to get the first sheet name
data = wb.worksheets[0].values
data = list(data)

#get the first column values
idx = [r[0] for r in data]
print(idx)

